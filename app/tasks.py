import openpyxl

from app.celery import app
from app.models import XlsxFile
from app.utils import get_pretty_list_from_string, get_rep_count, has_duplicates


@app.task(bind=True)
def xlsx_processing(self, pk: int):
    instance = XlsxFile.objects.get(pk=pk)
    workbook = openpyxl.load_workbook(filename=instance.file)

    before = None
    after = None

    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        max_column = worksheet.max_column
        if max_column != 1:
            for column in range(1, max_column + 1):
                title_cell = worksheet.cell(row=1, column=column)
                before = before or (title_cell.offset(row=1).value if title_cell.value == 'before' else None)
                after = after or (title_cell.offset(row=1).value if title_cell.value == 'after' else None)

                self.update_state(state='PROGRESS',
                                  meta={'current sheet': sheet, 'current column': column, 'total column': max_column})

                if before and after:
                    break

        if before and after:
            break

    if before and after:
        list_before = get_pretty_list_from_string(before)
        list_after = get_pretty_list_from_string(after)

        change_type = 'added' if len(list_before) < len(list_after) else 'removed'

        if has_duplicates(list_before) or has_duplicates(list_after):
            result = f'{change_type}: {next(iter(get_rep_count(list_before, list_after)))[0]}'
        else:
            result = f'{change_type}: {next(iter(set(list_before) ^ set(list_after)))}'
        return result

    return 'no result'
